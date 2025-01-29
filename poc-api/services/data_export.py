import os
from flask import current_app
from openpyxl import load_workbook

from ..helpers.enums import AttendanceStatusEnum

def __map_status_to_sheet_value(status):
    match status:
        case AttendanceStatusEnum.PRESENT.value:
            return 'X'
        case AttendanceStatusEnum.NOTIFIED_ABSENT.value:
            return 'VP'
        case _:
            return ''

def __populate_report_sheet(report_template_wb, worksheet_index, attendance_field_name, attendance_report_data):
    mass_work_sheet = report_template_wb.worksheets[worksheet_index]

    mass_work_sheet.cell(row=3, column=2, value=attendance_report_data.get('unit_id'))
    mass_work_sheet.cell(row=4, column=2, value=attendance_report_data.get('unit_name'))
    
    date_column_start_index = 7
    for a_sunday in attendance_report_data.get('sundays'):
        mass_work_sheet.cell(row=6, column=date_column_start_index, value=a_sunday.strftime('%d-%m-%Y'))
        date_column_start_index += 1

    attendance_row_start_index = 7

    for index, student in enumerate(attendance_report_data.get('attendance_data')):
        mass_work_sheet.cell(row=attendance_row_start_index, column=1, value=index+1)
        mass_work_sheet.cell(row=attendance_row_start_index, column=2, value=student.get('student_code'))
        mass_work_sheet.cell(row=attendance_row_start_index, column=3, value=student.get('saint_name'))
        mass_work_sheet.cell(row=attendance_row_start_index, column=4, value=student.get('last_name'))
        mass_work_sheet.cell(row=attendance_row_start_index, column=5, value=student.get('middle_name'))
        mass_work_sheet.cell(row=attendance_row_start_index, column=6, value=student.get('first_name'))

        attendance_column_start_index = 7
        for a_sunday in attendance_report_data.get('sundays'):
            student_attendance_entry = next((attendance for attendance in student.get('attendances') if attendance.get('attendance_date') == a_sunday), None)
            if (student_attendance_entry is not None):
                mass_work_sheet.cell(row=attendance_row_start_index, column=attendance_column_start_index, value=__map_status_to_sheet_value(student_attendance_entry.get(attendance_field_name)))
            attendance_column_start_index += 1

        attendance_row_start_index += 1

def export_unit_attendance_report_to_excel(attendance_report_data):
    template_path = os.path.join(current_app.root_path, 'templates', 'reports', 'attendance_export_template.xlsx')
    report_template_wb = load_workbook(template_path)
    
    __populate_report_sheet(report_template_wb, 0, 'mass_status', attendance_report_data)
    __populate_report_sheet(report_template_wb, 1, 'lesson_status', attendance_report_data)

    report_template_wb.save("/home/minhhikari/test_result.xlsx")